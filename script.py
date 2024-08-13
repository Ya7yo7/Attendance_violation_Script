from openpyxl import load_workbook, Workbook
import os
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta, time
from openpyxl import load_workbook
from datetime import datetime, time
import smtplib
 
from email.message import EmailMessage
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
 
import ssl #to keep internet connection secured
time_off_employees = []
all_employees = []
employee_ids = []
employees_dict = {}
accepted_execuses = ['Annual Vacation EG', 'Work From Home EG', 'Work From Sahel', 'Special occasion', 'Sick leave']
def read_ids():
    # Load the Excel file
    wb = load_workbook('Attendance.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue  # Skip empty rows 
        employee_ids.append(row[0])
class Employee:
    def __init__(self, name, execuse, id, sign_in_time, sign_out_time,sign_in_violation ,sign_out_violation , duration_violation, duration,N_1_email=None,violation_date=None):
        self.name = name
        self.execuse = execuse
        self.id = id
        self.sign_in_time = sign_in_time
        self.sign_out_time = sign_out_time
        self.sign_in_violation = sign_in_violation
        self.sign_out_violation = sign_out_violation
        self.duration_violation = duration_violation
        self.duration = duration
        self.N_1_email = N_1_email
        self.violation_date = violation_date
    def __repr__(self):
        return f"Employee({self.name}, {self.execuse}, {self.id}, {self.sign_in_time}, {self.sign_out_time}, {self.sign_in_violation}, {self.sign_out_violation}, {self.duration_violation}, {self.duration}, {self.N_1_email},{self.violation_date})"

def read_all_employees():
    # Load the Excel file
    wb = load_workbook('Attendance.xlsx')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue  # Skip empty rows 
        employee = Employee(row[1],None,row[0],row[6],row[9],False,False,'',0,row[14],row[5])
        if isinstance(employee.violation_date, str):
            employee.violation_date = None
        else:
            temp_date= datetime.strptime(str(employee.violation_date), '%Y-%m-%d %H:%M:%S')
            employee.violation_date = temp_date.date()
        all_employees.append(employee)
        employees_dict[employee.id] = employee

def read_employees_with_timeoff():
    # Load the Excel file
    wb = load_workbook('TimeOffDay.xlsx')
    sheet = wb['Sheet1']
    # Get the worksheet
    ws = wb.active
    # Read the data from the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue  # Skip empty rows
        employees_dict[row[2]].execuse = row[1]

def check_sign_in_violations():
    ten_oclock = time(10, 5, 0)
    one_oclock = time(13, 5, 0)
    for employee in employees_dict.values():
        if isinstance(employee.sign_in_time, str):
            if employee.sign_in_time == '#N/A':
                employee.sign_in_time = None
                if employee.execuse in accepted_execuses:
                    employee.sign_in_violation = False
                elif employee.execuse == None:
                    employee.sign_in_violation = True
                continue
            else:
                employee.sign_in_time = datetime.strptime(employee.sign_in_time, '%H:%M:%S')
        
        if employee.execuse in accepted_execuses:
            employee.sign_in_violation = False
            continue
            
        if employee.execuse == 'Half Day Work From Home EG':
            if employee.sign_in_time > one_oclock:
                employee.sign_in_violation = True
        elif employee.execuse == 'Half Day Vacation From Home EG':
            if employee.sign_in_time > ten_oclock:
                employee.sign_in_violation = True
        elif employee.execuse == None:
            if employee.sign_in_time > ten_oclock:
                employee.sign_in_violation = True
def check_sign_out_violations():
    four_oclock = time(15, 55, 0)
    one_oclock = time(12, 55, 0)

    for employee in employees_dict.values():
        if isinstance(employee.sign_out_time, str):
            if employee.sign_out_time == '#N/A':
                employee.sign_out_time = None
                if employee.execuse in accepted_execuses:
                    employee.sign_out_violation = False
                elif employee.execuse == None:
                    employee.sign_out_violation = True
                continue
            else:
                employee.sign_out_time = datetime.strptime(employee.sign_out_time, '%H:%M:%S')

        if employee.execuse in accepted_execuses:
            employee.sign_in_violation = False
            continue
       
        if employee.execuse == 'Half Day Work From Home EG':
            if employee.sign_out_time < four_oclock:
                employee.sign_out_violation = True
        elif employee.execuse == 'Half Day Vacation From Home EG':
            if employee.sign_out_time < one_oclock:
                employee.sign_out_violation = True
        elif employee.execuse == None:
            if employee.sign_out_time < four_oclock:
                employee.sign_out_violation = True
                
def NoShowNoExcuse():
    for employee in employees_dict.values():
        if employee.sign_in_time == None and employee.sign_out_time == None and not employee.execuse:
            print(f"Employee {employee.name} with ID {employee.id} is absent with no excuse.")        


def calculate_duration():
    for employee in employees_dict.values():
        
        if employee.sign_out_time is None or employee.sign_in_time is None:
            if employee.execuse in accepted_execuses:
                employee.duration_violation = 'No'
                continue
            employee.duration_violation = 'Yes'
            continue
        
        
        duration = (datetime.combine(datetime.min, employee.sign_out_time) - (datetime.combine(datetime.min, employee.sign_in_time)))
        duration_in_hours = duration.total_seconds() / 3600
        if duration_in_hours < 8:
            employee.duration_violation = 'Yes'
        elif 8 <= duration_in_hours < 10:
            employee.duration_violation = 'No'
        elif duration_in_hours >= 10:
            employee.duration_violation = 'over worked'
        employee.duration = duration_in_hours

def create_excel_for_n_1():
    for employee in employees_dict.values():
        if employee.N_1_email is not None:
            filename = employee.N_1_email + '.xlsx'
            # Check if the employee has any violations
            if employee.sign_in_violation or employee.sign_out_violation or employee.duration_violation == 'over worked' or employee.duration_violation == 'Yes':
                
                # Check if the file exists
                if os.path.exists(filename):
                    # Load existing workbook
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    # Create a new workbook
                    wb = Workbook()
                    ws = wb.active
                    # Write headers
                    headers = ["Employee Name", "ID", "Violation Type", "Violation Date"]
                    ws.append(headers)
                    for cell in ws[1]:  # ws[1] refers to the first row
                        cell.font = Font(bold=True)
                    

                # Determine the violation type(s)
                violation_types = []
                if employee.duration==0 and employee.execuse==None:
                    violation_types.append("Absent")
                else:
                    if employee.sign_in_violation:
                        violation_types.append("Sign-In Violation")
                    if employee.sign_out_violation:
                        violation_types.append("Sign-Out Violation")
                    if employee.duration_violation == 'Yes' and employee.execuse not in accepted_execuses:
                        violation_types.append("Duration Violation")
                    if employee.duration_violation == 'over worked':
                        violation_types.append("Overworked")
                

                # Add one row per violation type
                for violation in violation_types:
                    row = [employee.name, employee.id, violation, str(employee.violation_date)]
                    ws.append(row)
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2  # Add some extra padding

                # Save the workbook
                wb.save(filename)



 
 

 
def Send_Email(receiver_email,fileName):
    body = """
    Dear N+1,
    Kindly find the attached file for the attendance violations of your direct reports.
    Best Regards,
    HR Team
    """
    #mae MIME objec to define email parts
    
    smtp_port = 587
    smtp_server = "smtp.gmail.com"
    password = "fooj jzaa frvl xemo"
    sender_email = "mostafaamagdii7@gmail.com"
    subject = 'Attendance Violations'
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject
   
    #atttach body of the message
    msg.attach(MIMEText(body,'plain'))
   
 
   
    #open the file in python
    AttachedFile = open(fileName,'rb') #r for read b for binary
   
    #encoding the file in base 64
    attachedFile_package = MIMEBase('application','octet-stream')
    attachedFile_package.set_payload((AttachedFile).read())
    encoders.encode_base64(attachedFile_package)
    attachedFile_package.add_header('Content-Disposition',"AttachedFile ; filename = " + fileName)
    msg.attach(attachedFile_package)
   
   
    #cast as string
    text = msg.as_string()
    print("connecting to server")
    TIE_Server = smtplib.SMTP(smtp_server,smtp_port)
    TIE_Server.starttls()
    TIE_Server.login(sender_email,password)
   
    TIE_Server.sendmail(sender_email,receiver_email,text)
    print("Email Send Successfully")
   
    TIE_Server.quit()
   
def get_emails_from_filenames():
    # Get a list of all files in the current directory
    files = os.listdir('.')
    
    # Filter out the .xlsx files that contain '@' in the filename
    excel_files = [file for file in files if file.endswith('.xlsx') and '@' in file]
    
    # Extract emails by removing the .xlsx extension
    emails = [file.replace('.xlsx', '') for file in excel_files]
    
    return emails

# Call the function and store the emails in an array

                   
def main():
    read_ids()
    read_all_employees()
    read_employees_with_timeoff()
    check_sign_in_violations()
    check_sign_out_violations()
    NoShowNoExcuse()
    calculate_duration()
    for employee in employees_dict.values():
        print(employee,'\n')
    create_excel_for_n_1()
    email_list = get_emails_from_filenames()
    for email in email_list:
        Send_Email(email,email+'.xlsx')
if __name__ == '__main__':
    main()