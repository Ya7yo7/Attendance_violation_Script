from openpyxl import *
import os
import re
#############################################################################################################
def  Excel_create_or_load_excel(filename,new_sheet_name):
    if os.path.exists(filename):
        # Load existing workbook
        wb = load_workbook(filename)
    else:
        # Create new workbook
        wb = Workbook()
        # Save new workbook with the specified filename
        wb.save(filename)
        Excel_change_sheet_name(filename, 'Sheet',new_sheet_name)
    return wb
#############################################################################################################
def Excel_sheet_exists(filename):
    """Check if a sheet with the given name exists in the Excel file."""
    workbook = load_workbook(filename)
    return workbook.sheetnames
#############################################################################################################
def Excel_create_or_load_sheet(filename, sheet_name):
    """
    Create or load a sheet in an Excel file. If the file or sheet does not exist, create them.

    Args:
    filename (str): The name of the Excel file.
    sheet_name (str): The name of the sheet to create or load.
    
    Returns:
    openpyxl.worksheet.worksheet.Worksheet: The worksheet object.
    """
    # Check if the file exists
    if os.path.exists(filename):
        # Load the existing workbook
        wb = load_workbook(filename)
        # Check if the sheet exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    else:
        # Create a new workbook and add the sheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    
    # Save the workbook
    wb.save(filename)
    
    return ws

#############################################################################################################
def  Excel_change_sheet_name(filename, current_sheet_name, new_sheet_name):
    # Load the existing workbook
    wb = load_workbook(filename)

    # Check if the current sheet name exists
    if current_sheet_name in wb.sheetnames:
        # Change the sheet name
        ws = wb[current_sheet_name]
        ws.title = new_sheet_name

        # Save the modified workbook
        wb.save(filename)

#############################################################################################################
def  Excel_add_data_to_cell(filename, sheet_name, cell, data):
    # Load the existing workbook
    wb = load_workbook(filename)

    # Check if the sheet exists
    if sheet_name in wb.sheetnames:
        # Select the sheet
        ws = wb[sheet_name]

        # Add data to specific cell
        ws[cell] = data

        # Save the modified workbook
        wb.save(filename)
        
#############################################################################################################
def  Excel_split_cell_address(cell_address):
    """
    Split a cell address into column letters and row number.
    
    Args:
    cell_address (str): The cell address, e.g., 'B12'
    
    Returns:
    tuple: (column_letter(s), row_number)
    """
    # Use regular expressions to match the column letters and row number
    match = re.match(r"([A-Z]+)(\d+)", cell_address)
    
    if match:
        column_letters = match.group(1)
        row_number = int(match.group(2))
        return column_letters, row_number
    else:
        raise ValueError("Invalid cell address format")
#############################################################################################################
def  Excel_combine_column_row_CellAddress(column_letters, row_number):
    """
    Combine column letters and row number into a cell address.
    
    Args:
    column_letters (str): The column letters, e.g., 'B'.
    row_number (int): The row number, e.g., 12.
    
    Returns:
    str: The cell address, e.g., 'B12'.
    """
    return f'{column_letters}{row_number}'
#############################################################################################################
def  Excel_add_data_to_specific_cell(filename, sheet_name, column_letters, row_number, data):
    """
    Add data to a specific cell in an Excel sheet using column letters and row number.
    
    Args:
    filename (str): The Excel file to modify.
    sheet_name (str): The sheet in which to add the data.
    column_letters (str): The column letters of the cell, e.g., 'B'.
    row_number (int): The row number of the cell, e.g., 12.
    data (any): The data to insert into the cell.
    """
    # Load the existing workbook
    wb = load_workbook(filename)
    
    # Check if the sheet exists
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Combine column letters and row number into a cell address
    cell_address =  Excel_combine_column_row_CellAddress(column_letters, row_number)
    
    # Set the value in the specified cell
    ws[cell_address] = data
    
    # Save the modified workbook
    wb.save(filename)
#############################################################################################################
def Excel_clear_cell_data_from_specific_cell(filename, sheet_name,  column_letters, row_number):
    """
    Clear data from a specific cell in an Excel sheet.
    
    Args:
    filename (str): The Excel file to modify.
    sheet_name (str): The sheet in which to clear the data.
    cell_address (str): The address of the cell to clear, e.g., 'B12'.
    """
    filename = filename + '.xlsx'
    # Load the existing workbook
    wb = load_workbook(filename)
    cell_address =  Excel_combine_column_row_CellAddress(column_letters, row_number)

    # Check if the sheet exists
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file: {filename}")
    
    # Clear the data from the specified cell
    ws[cell_address] = None  # Setting the cell value to None clears it
    
    # Save the modified workbook
    wb.save(filename)
#############################################################################################################
